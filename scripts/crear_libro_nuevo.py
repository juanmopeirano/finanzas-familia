"""
crear_libro_nuevo.py — genera el nuevo libro 'Planilla Familia.xlsx' con la
estructura definitiva.

REGLAS:
  • 2025: se preserva tal cual el histórico (Categoría = Descripción vieja).
  • 2026+: se aplica el mapping nuevo (reglas + fallback CC).
  • Filas 'Saldo Inicial' / 'Saldo Final' viejas → se descartan (ya no se usan).
  • Se conservan TODAS las columnas operativas: Tipo, Origen, Categoría,
    Subcategoría, Fecha, Concepto, Débito, Crédito, Saldo, Referencia, Destino.
  • Se agrega columna 'ML_Revisar' (para marcar las que ML dudó).
  • Dropdowns en Categoría / Subcategoría (no se puede inventar).
  • Hoja 'Categorías' con la referencia.

Uso: python crear_libro_nuevo.py
"""

import sys
import shutil
import tempfile
import os
import warnings
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).parent))
from config import EXCEL_VIEJO as EXCEL_PATH, SHEET_CA_VIEJO as SHEET_CA, SHEET_CC_VIEJO as SHEET_CC
from reglas import clasificar, cargar_reglas
from generar_mapping import CATEGORIAS
from export_json import CC_FALLBACK_OLD_CAT, CAT_VIEJA_NUEVA

warnings.filterwarnings("ignore", message="Data Validation extension is not supported")
warnings.filterwarnings("ignore", category=UserWarning)


# ── configuración ────────────────────────────────────────────────────────────
from config import EXCEL_PATH as EXCEL_NUEVO  # leer destino del config
SHEET_CA_NUEVO    = "CA - Caja de Ahorro"
SHEET_CC_NUEVO    = "CC - Cuenta Corriente"
ANO_INICIO_NUEVO  = 2026  # 2026+ = mapping nuevo, 2025 = preservar

COLUMNAS = ["Tipo", "Origen", "Categoría", "Subcategoría", "Fecha",
            "Concepto", "Débito", "Crédito", "Saldo", "Referencia",
            "Destino", "ML_Revisar"]

ANCHOS = {"Tipo": 12, "Origen": 18, "Categoría": 24, "Subcategoría": 22,
          "Fecha": 12, "Concepto": 38, "Débito": 12, "Crédito": 12,
          "Saldo": 13, "Referencia": 14, "Destino": 14, "ML_Revisar": 11}

# colores
HEADER_FILL = PatternFill("solid", fgColor="064E3B")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WARN_FILL   = PatternFill("solid", fgColor="FEF3C7")  # amarillo


# ── lectura del Excel viejo ──────────────────────────────────────────────────

def read_excel_safe(path, sheet):
    try:
        return pd.read_excel(path, sheet_name=sheet, engine="openpyxl")
    except PermissionError:
        tmp = os.path.join(tempfile.gettempdir(), f"_temp_{Path(path).name}")
        shutil.copy2(path, tmp)
        return pd.read_excel(tmp, sheet_name=sheet, engine="openpyxl")


# ── transformar filas (vieja → nueva) ────────────────────────────────────────

def transformar(df, aplicar_mapping_nuevo=True):
    """
    Si aplicar_mapping_nuevo=True (CA): 2026+ usa reglas nuevas, 2025 preserva.
    Si aplicar_mapping_nuevo=False (CC): SIEMPRE preserva el histórico (manual).
    """
    cargar_reglas()
    out = []
    for _, r in df.iterrows():
        fecha = r.get("Fecha")
        try:
            fecha = pd.to_datetime(fecha, errors="coerce", dayfirst=True)
        except Exception:
            continue
        if pd.isna(fecha) or fecha.year < 2020:
            continue

        tipo = str(r.get("Tipo", "") or "").strip()
        if tipo in ("Saldo Inicial", "Saldo Final"):
            continue

        descripcion  = str(r.get("Descripción", "") or "").strip()
        descripcion2 = str(r.get("Descripción 2", "") or "").strip()
        concepto     = str(r.get("Concepto", "") or "").strip()

        if aplicar_mapping_nuevo and fecha.year >= ANO_INICIO_NUEVO:
            # CA 2026+ — PRIORIZAR label manual viejo (traducido al schema nuevo)
            # Si no hay label, aplicar reglas (caso de imports nuevos)
            if descripcion and descripcion in CAT_VIEJA_NUEVA:
                cat, ns = CAT_VIEJA_NUEVA[descripcion]
                sub = ns or descripcion2  # usar sub viejo si no hay default
                ml_revisar = False
            elif descripcion:
                # Hay label manual pero no está en el mapeo - dejarlo igual
                cat = descripcion
                sub = descripcion2
                ml_revisar = False
            else:
                # Sin label manual → aplicar reglas
                cat, sub = clasificar(concepto)
                ml_revisar = (cat == "Sin clasificar")
        else:
            # CA 2025 / CC todos — preservar tal cual
            cat = descripcion
            sub = descripcion2
            ml_revisar = False

        def num(x):
            if pd.isna(x): return None
            try:    return float(x)
            except: return None

        out.append({
            "Tipo":         tipo,
            "Origen":       str(r.get("Origen", "") or "").strip(),
            "Categoría":    cat,
            "Subcategoría": sub,
            "Fecha":        fecha.date(),
            "Concepto":     concepto,
            "Débito":       num(r.get("Débito")),
            "Crédito":      num(r.get("Crédito")),
            "Saldo":        num(r.get("Saldo")),
            "Referencia":   str(r.get("Referencia", "") or "").strip(),
            "Destino":      str(r.get("Destino", "") or "").strip(),
            "ML_Revisar":   ml_revisar,
        })
    return out


# ── escritura del libro nuevo ────────────────────────────────────────────────

def crear_hoja_categorias(wb):
    ws = wb.create_sheet("Categorías")

    # Tabla referencia categoría → subcategorías
    ws.cell(1, 1, "Categoría").font = HEADER_FONT
    ws.cell(1, 1).fill = HEADER_FILL
    ws.cell(1, 2, "Subcategorías permitidas").font = HEADER_FONT
    ws.cell(1, 2).fill = HEADER_FILL

    cats_list = list(CATEGORIAS.keys())
    for i, (cat, subs) in enumerate(CATEGORIAS.items(), 2):
        ws.cell(i, 1, cat).font = Font(bold=True)
        ws.cell(i, 2, " · ".join(subs) if subs else "(sin subcategorías)")

    # Columna oculta D = lista plana de TODAS las subcategorías (para dropdown)
    ws.cell(1, 4, "_Subcategorías_lista").font = HEADER_FONT
    ws.cell(1, 4).fill = HEADER_FILL
    subs_flat = sorted(set(s for subs in CATEGORIAS.values() for s in subs))
    for i, s in enumerate(subs_flat, 2):
        ws.cell(i, 4, s)
    # se agrega "" como opción para subcat vacía
    ws.cell(len(subs_flat) + 2, 4, "(vacía)")

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["D"].width = 24
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    return len(cats_list), len(subs_flat) + 1  # +1 por la opción "(vacía)"


def escribir_hoja_movs(ws, rows, n_cats, n_subs, con_validacion=True):
    # Headers
    for c, h in enumerate(COLUMNAS, 1):
        cell = ws.cell(1, c, h)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Data
    for i, r in enumerate(rows, 2):
        for c, col in enumerate(COLUMNAS, 1):
            val = r.get(col)
            if val == "" or val is None:
                val = None
            ws.cell(i, c, val)
        # Marcar amarillo si requiere revisión
        if r.get("ML_Revisar"):
            for col_name in ("Categoría", "Subcategoría"):
                ws.cell(i, COLUMNAS.index(col_name) + 1).fill = WARN_FILL

    # Anchos + formatos
    for c, h in enumerate(COLUMNAS, 1):
        ws.column_dimensions[get_column_letter(c)].width = ANCHOS.get(h, 12)

    fc = COLUMNAS.index("Fecha") + 1
    for i in range(2, len(rows) + 2):
        ws.cell(i, fc).number_format = "DD/MM/YYYY"
    for col_name in ("Débito", "Crédito", "Saldo"):
        cn = COLUMNAS.index(col_name) + 1
        for i in range(2, len(rows) + 2):
            ws.cell(i, cn).number_format = "#,##0.00"

    # Data validation (solo en CA - en CC el usuario maneja manualmente)
    if con_validacion:
        cat_letter = get_column_letter(COLUMNAS.index("Categoría") + 1)
        sub_letter = get_column_letter(COLUMNAS.index("Subcategoría") + 1)
        range_max  = max(len(rows) + 200, 3000)

        dv_cat = DataValidation(
            type="list",
            formula1=f"=Categorías!$A$2:$A${n_cats + 1}",
            allow_blank=True,
            showDropDown=False,
        )
        dv_cat.error = "Elegí una categoría de la lista."
        dv_cat.errorTitle = "Categoría inválida"
        ws.add_data_validation(dv_cat)
        dv_cat.add(f"{cat_letter}2:{cat_letter}{range_max}")

        dv_sub = DataValidation(
            type="list",
            formula1=f"=Categorías!$D$2:$D${n_subs + 1}",
            allow_blank=True,
            showDropDown=False,
        )
        ws.add_data_validation(dv_sub)
        dv_sub.add(f"{sub_letter}2:{sub_letter}{range_max}")

    ws.freeze_panes = "C2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNAS))}{len(rows) + 1}"


def crear_libro():
    wb = Workbook()
    # quitar la hoja por defecto
    wb.remove(wb.active)

    # 1) Hoja referencia (primero, para que las dropdowns la encuentren)
    n_cats, n_subs = crear_hoja_categorias(wb)

    # 2) Hoja CA — usa mapping nuevo desde 2026
    print("  Procesando CA...")
    ca_old  = read_excel_safe(EXCEL_PATH, SHEET_CA)
    ca_rows = transformar(ca_old, aplicar_mapping_nuevo=True)
    ws_ca   = wb.create_sheet(SHEET_CA_NUEVO, 0)
    escribir_hoja_movs(ws_ca, ca_rows, n_cats, n_subs, con_validacion=True)
    print(f"    {len(ca_rows)} filas")

    # 3) Hoja CC — preserva el histórico siempre (manual del usuario)
    print("  Procesando CC...")
    cc_old  = read_excel_safe(EXCEL_PATH, SHEET_CC)
    cc_rows = transformar(cc_old, aplicar_mapping_nuevo=False)
    ws_cc   = wb.create_sheet(SHEET_CC_NUEVO, 1)
    escribir_hoja_movs(ws_cc, cc_rows, n_cats, n_subs, con_validacion=False)
    print(f"    {len(cc_rows)} filas")

    # Guardar
    wb.save(EXCEL_NUEVO)

    # Stats
    revisar_ca = sum(1 for r in ca_rows if r["ML_Revisar"])
    revisar_cc = sum(1 for r in cc_rows if r["ML_Revisar"])
    sin_2026_ca = sum(1 for r in ca_rows if r["Fecha"].year >= 2026)
    sin_2026_cc = sum(1 for r in cc_rows if r["Fecha"].year >= 2026)
    return EXCEL_NUEVO, ca_rows, cc_rows, revisar_ca, revisar_cc


def main():
    print(f"Cargando datos de: {EXCEL_PATH.name}")
    cargar_reglas()
    print(f"Generando libro nuevo...")
    out, ca, cc, rca, rcc = crear_libro()
    try:
        print(f"\nOK -> {out}")
        print(f"  CA: {len(ca)} filas, {rca} para revisar (amarillo)")
        print(f"  CC: {len(cc)} filas, {rcc} para revisar (amarillo)")
    except UnicodeEncodeError:
        print(f"\nOK ({len(ca)} CA, {len(cc)} CC, {rca+rcc} para revisar)")


if __name__ == "__main__":
    main()
