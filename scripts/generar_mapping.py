"""
generar_mapping.py — analiza los Conceptos del banco en el Excel maestro
y genera 'mapping_propuesto.xlsx' con reglas para revisar manualmente.

3 hojas:
  1. Patrones — reglas detectadas con confianza, para confirmar/editar
  2. Sin patrón — Conceptos únicos que no matchearon ninguna regla
  3. Categorías — referencia de categorías/subcategorías válidas
"""

import sys
import re
from collections import Counter, defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).parent))
from config import EXCEL_PATH, SHEET_CA, PROJECT_DIR

OUTPUT = PROJECT_DIR / "mapping_propuesto.xlsx"

# ── árbol de categorías nuevo ──────────────────────────────────────────────────
CATEGORIAS = {
    "Supermercado":        ["Tienda Inglesa", "Disco", "Devoto", "Geant", "Macromercado",
                            "Kinko", "Binmon", "Frog", "La Molienda", "Palermo", "Otro"],
    "Servicios":           ["ANTEL", "UTE", "OSE", "Personal Bank", "CADE", "Otro"],
    "Cuotas":              ["Honda", "Celerio", "Otro"],
    "Gastos comunes":      [],
    "Niñera":              [],
    "Jardín":              [],
    "Seguro de vida":      [],
    "Mapfre":              ["Casa", "Auto", "Otro"],
    "CJPPU":               [],
    "Salud":               ["Farmacia", "Dentista", "BCBS", "Médico", "Estudios", "Otro"],
    "Estación de servicio": ["Combustible", "Peaje", "Otro"],
    "Comida trabajo":      [],
    "Social / Amigos":     ["Restaurant", "Cine/Teatro", "Paseo", "Café", "Cabify/Uber",
                            "Juntada", "Cumpleaños", "Cena grupal", "Otro"],
    "Regalos":             ["Cumple", "Navidad", "Bebé", "Casamiento", "Otro"],
    "Delivery / Pedidos":  ["Pedidos Ya", "Rappi", "Reintegro amigos", "Otro"],
    "Entretenimiento":     ["Streaming", "Evento/Show", "Hobbies", "Otro"],
    "Deportes / Gym":      ["Gym", "Whoop", "Equipamiento", "Otro"],
    "Viajes":              ["Pasajes", "Alojamiento", "En destino", "Otro"],
    "Hogar - Mejoras":     ["Muebles", "Decoración", "Reparaciones", "Electrodomésticos",
                            "Útiles", "Ferretería", "Otro"],
    "Ropa":                ["JM", "Pili", "Otro"],
    "Cosmética":           [],
    "Varios":              [],
    "Sueldo JM":           [],
    "Sueldo Pili":         [],
    "Otros ingresos":      ["Reintegro", "Bonus", "Regalo", "Otro"],
    "Ahorros":             [],
    "No va":               [],
    "Traspaso":            [],
    "Sin clasificar":      [],
}

# ── mapeo categoría vieja → categoría nueva (para autorellenar Sin patrón) ──
CAT_VIEJA_NUEVA = {
    "Supermercado":                 ("Supermercado",          ""),
    "Ocio, social, entrenimento":   ("Social / Amigos",       ""),
    "Extras Casa":                  ("Hogar - Mejoras",       ""),
    "No va":                        ("No va",                 ""),
    "Farmacia":                     ("Salud",                 "Farmacia"),
    "Comida oficina":               ("Comida trabajo",        ""),
    "Nafta":                        ("Estación de servicio",  "Combustible"),
    "Paquete Personal Bank":        ("Servicios",             "Personal Bank"),
    "ANTEL":                        ("Servicios",             "ANTEL"),
    "Otros":                        ("Varios",                ""),
    "Caja de Ahorro":               ("No va",                 ""),
    "Cuenta Corriente":             ("No va",                 ""),
    "Salud (Dentista, BCBS, ots)":  ("Salud",                 "Dentista"),
    "Limpieza Casa":                ("Niñera",                ""),
    "Sueldo Pili":                  ("Sueldo Pili",           ""),
    "Sueldo":                       ("Sueldo JM",             ""),
    "Honda":                        ("Cuotas",                "Honda"),
    "CJPPU":                        ("CJPPU",                 ""),
    "Celerio":                      ("Cuotas",                "Celerio"),
    "Seguro de vida":               ("Seguro de vida",        ""),
    "UTE":                          ("Servicios",             "UTE"),
    "OSE":                          ("Servicios",             "OSE"),
    "Gastos comunes":               ("Gastos comunes",        ""),
    "Jardinero":                    ("Jardín",                ""),
    "Cosmética":                    ("Cosmética",             ""),
    "CADE":                         ("Servicios",             "CADE"),
    "Traspaso a Santander Pili":    ("Traspaso",              ""),
    "Traspaso de Santander Pili":   ("Traspaso",              ""),
    "Traspaso de Itau Pili":        ("Traspaso",              ""),
    "Traspaso de CC":               ("Traspaso",              ""),
    "Ahorros":                      ("Ahorros",               ""),
}

# ── traducción de mi propuesta para cada patrón conocido ──────────────────────
# Esto es lo que VOS vas a revisar y corregir en el Excel.
PROPUESTA = {
    # Supermercados
    "TIENDA":       ("Supermercado", "Tienda Inglesa", "Tienda Inglesa"),
    "INGLE":        ("Supermercado", "Tienda Inglesa", "Tienda Inglesa"),
    "INLGE":        ("Supermercado", "Tienda Inglesa", "(typo de Tienda Inglesa)"),
    "MACROMERCADO": ("Supermercado", "Macromercado", ""),
    "KINKO":        ("Supermercado", "Kinko", ""),
    "BINMON":       ("Supermercado", "Binmon", ""),
    "DEVOTO":       ("Supermercado", "Devoto", ""),
    "EXPRE":        ("Supermercado", "Devoto", "Devoto Express"),
    "FROG":         ("Supermercado", "Frog", ""),
    "MOLIENDA":     ("Supermercado", "La Molienda", ""),
    "PALERMO":      ("Supermercado", "Palermo", "Palermo Hnos"),
    "GEANT":        ("Supermercado", "Geant", ""),
    "QUESO":        ("Supermercado", "Otro", "Queso Bus"),
    "BUS":          ("Supermercado", "Otro", "Queso Bus"),

    # Servicios
    "ANTEL":        ("Servicios", "ANTEL", ""),
    "FIJO":         ("Servicios", "ANTEL", "ANTEL Fijo"),
    "MOVIL":        ("Servicios", "ANTEL", "ANTEL Móvil"),
    "AUT":          ("Servicios", "ANTEL", "(débito automático ANTEL)"),
    "UTE":          ("Servicios", "UTE", ""),
    "OSE":          ("Servicios", "OSE", ""),
    "PAQUETE":      ("Servicios", "Personal Bank", "Paquete Personal Bank"),

    # Salud
    "FARMACIA":     ("Salud", "Farmacia", ""),
    "FARMASHOP":    ("Salud", "Farmacia", ""),
    "ROQUE":        ("Salud", "Farmacia", "San Roque"),

    # Hogar
    "SODIMAC":      ("Hogar - Mejoras", "Ferretería", ""),
    "FERRETERIA":   ("Hogar - Mejoras", "Ferretería", ""),
    "ACHER":        ("Hogar - Mejoras", "Ferretería", "Acher Cerámica"),
    "CERAMICA":     ("Hogar - Mejoras", "Ferretería", "Acher Cerámica"),
    "KROSER":       ("Hogar - Mejoras", "Ferretería", ""),
    "UNISP":        ("Hogar - Mejoras", "Ferretería", "(parte de Kroser?)"),

    # Cosmética / belleza
    "DEPILIFE":     ("Cosmética", "", ""),
    "CALORSLIM":    ("Cosmética", "", "¿es estética/cuidado personal?"),

    # Delivery
    "DLO":          ("Delivery / Pedidos", "Pedidos Ya", "DLO = Pedidos Ya"),
    "PEDIDOSY":     ("Delivery / Pedidos", "Pedidos Ya", ""),
    "PEDIDOSYA":    ("Delivery / Pedidos", "Pedidos Ya", ""),

    # Entretenimiento
    "GOOGLE":       ("Entretenimiento", "Hobbies", "apps/juegos Google Play"),
    "SLAYER":       ("Entretenimiento", "Hobbies", "juego mobile"),
    "LEGEN":        ("Entretenimiento", "Hobbies", "juego mobile (Slayer Legends)"),
    "MOVIE":        ("Social / Amigos", "Cine/Teatro", ""),

    # Auto/Transporte
    "ESTACION":     ("Estación de servicio", "Combustible", ""),
    "ITA":          ("Estación de servicio", "Combustible", "Estación Itaú"),
    "TELEPEAJE":    ("Estación de servicio", "Peaje", ""),
    "CABI":         ("Social / Amigos", "Cabify/Uber", "Cabify"),

    # Otros
    "VIDA":         ("Seguro de vida", "", ""),
    "SEGURO":       ("Seguro de vida", "", ""),
    "SOBRE":        ("Seguro de vida", "", "(\"sobre saldo\")"),
    "APORTES":      ("CJPPU", "", ""),
    "ADMREN":       ("Gastos comunes", "", ""),
    "RUDY":         ("Social / Amigos", "Restaurant", "¿Rudy es un bar/restaurant?"),
    "SUPER":        ("Supermercado", "Otro", ""),
}


# ── análisis de patrones (mismo algoritmo que antes) ──────────────────────────
STOP = {"COMPRA","REDIVA","PAGO","TRANSFERENCIA","DEBITO","CREDITO","POS","WEB",
        "TARJETA","ITAU","CAJA","DE","AHORRO","CRE","DEB","CAMBIOSST","CAMBIOSOP",
        "VARIOS","VISA","LINK","ILINK","TRASPASO","DESDE","HASTA","CUENTA","A","EN","LA","EL"}


def norm(s):
    s = str(s).upper()
    s = re.sub(r"[\.\*\-/]", " ", s)
    s = re.sub(r"\d+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def tokens(s):
    return [t for t in s.split() if t not in STOP and len(t) > 2]


def detectar_patrones(df):
    df["cn"]   = df["Concepto"].apply(norm)
    df["toks"] = df["cn"].apply(tokens)

    token_cat = defaultdict(Counter)
    for _, r in df.iterrows():
        for t in set(r["toks"]):
            token_cat[t][r["Descripción"]] += 1

    patrones = []
    for tok, cats in token_cat.items():
        total = sum(cats.values())
        if total < 3: continue
        dom_cat, dom_n = cats.most_common(1)[0]
        if dom_n / total >= 0.85:
            patrones.append({
                "patron":      tok,
                "cat_vieja":   dom_cat,
                "matches":     dom_n,
                "total_apar":  total,
                "pureza":      round(dom_n / total, 2),
            })
    patrones.sort(key=lambda x: -x["matches"])
    return patrones, df


def conceptos_sin_patron(df, patrones):
    pats = {p["patron"] for p in patrones}
    df["matched"] = df["toks"].apply(lambda toks: any(t in pats for t in toks))
    sin_match = df[~df["matched"]].copy()

    # agrupar por concepto normalizado
    grouped = sin_match.groupby("cn").agg(
        apariciones = ("cn", "size"),
        cat_vieja_top = ("Descripción", lambda s: s.mode().iloc[0] if len(s) else ""),
        ejemplo = ("Concepto", "first"),
    ).reset_index().sort_values("apariciones", ascending=False)
    return grouped


# ── escritura del Excel ──────────────────────────────────────────────────────
HEADER_FILL    = PatternFill("solid", fgColor="064E3B")
HEADER_FONT    = Font(color="FFFFFF", bold=True, size=11)
SUGERIDO_FILL  = PatternFill("solid", fgColor="FEF3C7")  # amarillo claro
NOTA_FILL      = PatternFill("solid", fgColor="EFF6FF")  # azul claro
SECTION_FILL   = PatternFill("solid", fgColor="065F46")
THIN           = Side(border_style="thin", color="E5E7EB")
BOX            = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)


def escribir_excel(patrones, sin_pat, output=OUTPUT):
    wb = Workbook()

    # ─── Sheet 1: Patrones ─────────────────────────────────
    ws = wb.active
    ws.title = "Patrones"

    headers = ["#", "Patrón (token banco)", "Categoría", "Subcategoría", "Categoría vieja",
               "# matches", "Pureza", "Ejemplo concepto", "Nota / sugerencia"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    df_origen = _read_excel_safe(EXCEL_PATH, SHEET_CA)
    df_origen["cn"]   = df_origen["Concepto"].fillna("").astype(str).apply(norm)

    for i, p in enumerate(patrones, 1):
        prop = PROPUESTA.get(p["patron"], ("", "", ""))
        cat, sub, nota = prop
        ej_row = df_origen[df_origen["cn"].str.contains(p["patron"], na=False, regex=False)]
        ejemplo = str(ej_row["Concepto"].iloc[0])[:55] if len(ej_row) else ""

        row = [i + 1, p["patron"], cat, sub, p["cat_vieja"],
               p["matches"], p["pureza"], ejemplo, nota]
        for c, v in enumerate(row, 1):
            cell = ws.cell(i + 1, c, v)
            cell.border = BOX
            cell.alignment = Alignment(vertical="center", wrap_text=(c == 8))
            if c in (3, 4):  # categoría / subcat
                cell.fill = SUGERIDO_FILL
                cell.font = Font(bold=True)
            if c == 9:
                cell.fill = NOTA_FILL
                cell.font = Font(italic=True, color="475569")

    # anchos
    widths = [4, 22, 22, 22, 28, 10, 8, 38, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # validación de datos para Categoría
    cats_str = ",".join(CATEGORIAS.keys())
    if len(cats_str) < 250:
        dv_cat = DataValidation(type="list", formula1=f'"{cats_str}"', allow_blank=True)
        dv_cat.error = "Elegí una de las categorías válidas"
        ws.add_data_validation(dv_cat)
        dv_cat.add(f"C2:C{len(patrones) + 50}")
    else:
        # demasiado largo para inline list — usamos referencia a hoja Categorías
        dv_cat = DataValidation(type="list", formula1=f"=Categorías!$A$2:$A${len(CATEGORIAS) + 1}",
                                allow_blank=True)
        ws.add_data_validation(dv_cat)
        dv_cat.add(f"C2:C{len(patrones) + 50}")

    # ─── Sheet 2: Conceptos sin patrón ─────────────────────────────────
    ws2 = wb.create_sheet("Sin patrón")
    headers2 = ["#", "Concepto normalizado", "# apariciones", "Cat. vieja más común",
                "Categoría a asignar", "Subcategoría", "Ejemplo original"]
    for c, h in enumerate(headers2, 1):
        cell = ws2.cell(1, c, h)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, (_, r) in enumerate(sin_pat.iterrows(), 1):
        cat_nueva, sub_nueva = CAT_VIEJA_NUEVA.get(r["cat_vieja_top"], ("", ""))
        row = [i, r["cn"][:60], int(r["apariciones"]), r["cat_vieja_top"],
               cat_nueva, sub_nueva, str(r["ejemplo"])[:55]]
        for c, v in enumerate(row, 1):
            cell = ws2.cell(i + 1, c, v)
            cell.border = BOX
            cell.alignment = Alignment(vertical="center", wrap_text=(c in (2, 7)))
            if c in (5, 6):
                cell.fill = SUGERIDO_FILL

    widths2 = [5, 45, 12, 25, 22, 22, 38]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.row_dimensions[1].height = 28
    ws2.freeze_panes = "A2"

    # validación categoría sheet 2
    dv_cat2 = DataValidation(type="list",
                             formula1=f"=Categorías!$A$2:$A${len(CATEGORIAS) + 1}",
                             allow_blank=True)
    ws2.add_data_validation(dv_cat2)
    dv_cat2.add(f"E2:E{len(sin_pat) + 50}")

    # ─── Sheet 3: Categorías (referencia) ─────────────────────────────────
    ws3 = wb.create_sheet("Categorías")
    ws3.cell(1, 1, "Categoría").fill = HEADER_FILL
    ws3.cell(1, 1).font = HEADER_FONT
    ws3.cell(1, 2, "Subcategorías válidas").fill = HEADER_FILL
    ws3.cell(1, 2).font = HEADER_FONT
    for r, (cat, subs) in enumerate(CATEGORIAS.items(), 2):
        ws3.cell(r, 1, cat).font = Font(bold=True)
        ws3.cell(r, 2, " · ".join(subs) if subs else "(sin subcategorías)")
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 80
    ws3.row_dimensions[1].height = 26
    ws3.freeze_panes = "A2"

    # ─── Sheet 4: Instrucciones ─────────────────────────────────
    ws4 = wb.create_sheet("Instrucciones", 0)  # primera hoja
    instr = [
        ("FINANZAS — MAPPING PROPUESTO", True),
        ("", False),
        ("Generado a partir de los 2.747 movimientos del Excel maestro.", False),
        ("", False),
        ("HOJAS:", True),
        ("  • Patrones — reglas detectadas con alta confianza (>85% pureza)", False),
        ("  • Sin patrón — conceptos que NO matchean ninguna regla, hay que asignarlos", False),
        ("  • Categorías — referencia con todas las categorías y subcategorías válidas", False),
        ("", False),
        ("CÓMO USAR:", True),
        ("  1. Andá a la hoja 'Patrones'.", False),
        ("  2. Para cada fila revisá las columnas amarillas (Categoría / Subcategoría):", False),
        ("     • Si la sugerencia está bien → no toques nada", False),
        ("     • Si está mal → escribí la categoría correcta (usá el dropdown)", False),
        ("     • Si NO querés esa regla → borrá la fila completa", False),
        ("  3. Mirá la columna 'Nota' — ahí dejé observaciones donde dudo", False),
        ("  4. Después andá a 'Sin patrón' y completá las que reconozcas", False),
        ("  5. Lo que dejes vacío → se clasificará como 'Sin clasificar' y lo hacés a mano", False),
        ("", False),
        ("DEVOLVER:", True),
        ("  Guardá el archivo y avisame que ya lo revisaste.", False),
        ("  Yo lo proceso y construyo el clasificador final basado en estas reglas.", False),
    ]
    for r, (txt, is_header) in enumerate(instr, 1):
        cell = ws4.cell(r, 1, txt)
        if is_header:
            cell.font = Font(bold=True, size=13, color="064E3B")
        else:
            cell.font = Font(size=11, color="1F2937")
    ws4.column_dimensions["A"].width = 90

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


# ── main ─────────────────────────────────────────────────────────────────────
def _read_excel_safe(path, sheet):
    """Si el archivo está bloqueado por Excel/OneDrive, copia a temp y lee."""
    import shutil, tempfile, os
    try:
        return pd.read_excel(path, sheet_name=sheet, engine="openpyxl")
    except PermissionError:
        tmp = os.path.join(tempfile.gettempdir(), "planilla_temp.xlsx")
        shutil.copy2(path, tmp)
        return pd.read_excel(tmp, sheet_name=sheet, engine="openpyxl")


def main():
    print("Leyendo Excel maestro...")
    df = _read_excel_safe(EXCEL_PATH, SHEET_CA)
    df["Descripción"] = df["Descripción"].fillna("").astype(str).str.strip()
    df["Concepto"]    = df["Concepto"].fillna("").astype(str).str.strip()
    df = df[df["Descripción"] != ""].copy()
    print(f"  {len(df)} movimientos etiquetados")

    print("Detectando patrones...")
    patrones, df = detectar_patrones(df)
    print(f"  {len(patrones)} patrones encontrados (pureza >= 85%, >= 3 apariciones)")

    print("Buscando conceptos sin patrón...")
    sin_pat = conceptos_sin_patron(df, patrones)
    print(f"  {len(sin_pat)} conceptos únicos sin patrón")

    print(f"Generando {OUTPUT.name}...")
    out = escribir_excel(patrones, sin_pat)
    try:
        print(f"OK — archivo: {out}")
    except UnicodeEncodeError:
        print(f"OK ({len(patrones)} patrones, {len(sin_pat)} sin patron)")


if __name__ == "__main__":
    main()
