"""
classify.py — entrena un clasificador ML sobre el historial etiquetado
y sugiere categorías para movimientos sin clasificar.

Uso directo: python classify.py
"""

import sys
import re
import numpy as np
import pandas as pd
from pathlib import Path
from collections import Counter

sys.path.insert(0, str(Path(__file__).parent))
from config import EXCEL_PATH, SHEET_CA, CONFIANZA_MINIMA
from categorias import MAPEO, EXCLUIR_DESC, EXCLUIR_TIPO, STOP_BANCO

try:
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.compose import ColumnTransformer
    from sklearn.pipeline import Pipeline
    from sklearn.preprocessing import StandardScaler
    from sklearn.svm import LinearSVC
    from sklearn.calibration import CalibratedClassifierCV
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font
    HAS_SKLEARN = True
except ImportError:
    HAS_SKLEARN = False


# ── helpers ──────────────────────────────────────────────────────────────────

def to_number(x):
    if pd.isna(x): return 0.0
    if isinstance(x, (int, float, np.number)): return float(x)
    s = str(x).strip().replace(".", "").replace(",", ".")
    try: return float(s)
    except: return 0.0


def norm_text(s):
    s = "" if pd.isna(s) else str(s).upper()
    for ch in "*-/.": s = s.replace(ch, " ")
    s = re.sub(r"\d+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    toks = [t for t in s.split() if t not in STOP_BANCO and len(t) > 2]
    return " ".join(toks)


def apply_mapeo(df):
    """Migra categorías viejas al esquema nuevo in-place."""
    for old, (new, sub) in MAPEO.items():
        mask = df["Descripción"] == old
        if not mask.any():
            continue
        df.loc[mask, "Descripción"] = new
        if sub:
            no_sub = mask & (df["Descripción 2"].fillna("").str.strip() == "")
            df.loc[no_sub, "Descripción 2"] = sub
    return df


def load_sheet(excel_path=EXCEL_PATH, sheet_name=SHEET_CA):
    df = pd.read_excel(excel_path, sheet_name=sheet_name, engine="openpyxl")
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce", dayfirst=True)
    for col in ["Débito", "Crédito"]:
        df[col] = df[col].apply(to_number)
    df["Monto"] = df["Crédito"] - df["Débito"]
    for c in ["Concepto", "Referencia", "Destino", "Origen"]:
        if c not in df.columns:
            df[c] = ""
    df["TextoML"] = (
        df[["Concepto", "Referencia", "Destino", "Origen"]]
        .fillna("").astype(str).agg(" ".join, axis=1)
        .apply(norm_text)
    )
    df["Descripción"]   = df["Descripción"].fillna("").astype(str).str.strip()
    df["Descripción 2"] = df["Descripción 2"].fillna("").astype(str).str.strip()
    df["ML_Revisar"]    = df.get("ML_Revisar", pd.Series(False, index=df.index)).fillna(False)
    df["ML_Confianza"]  = pd.to_numeric(df.get("ML_Confianza", pd.Series(np.nan, index=df.index)), errors="coerce")
    apply_mapeo(df)
    return df


# ── entrenamiento ─────────────────────────────────────────────────────────────

def build_pipeline(train_df):
    counts   = Counter(train_df["Descripción"])
    min_cls  = min(counts.values())
    cv_folds = 2 if min_cls >= 2 else 1

    pre = ColumnTransformer([
        ("txt", TfidfVectorizer(ngram_range=(1, 2), min_df=1, sublinear_tf=True), "TextoML"),
        ("amt", Pipeline([("sc", StandardScaler())]), ["Monto"]),
    ])
    svc = LinearSVC(max_iter=3000, C=1.0)
    clf = CalibratedClassifierCV(svc, method="sigmoid", cv=cv_folds) if cv_folds >= 2 else svc
    pipe = Pipeline([("pre", pre), ("clf", clf)])
    pipe.fit(train_df[["TextoML", "Monto"]], train_df["Descripción"])
    return pipe


def predict_proba_safe(pipe, X):
    clf = pipe.named_steps["clf"]
    if hasattr(clf, "predict_proba"):
        proba   = pipe.predict_proba(X)
        classes = clf.classes_
        idx     = np.argmax(proba, axis=1)
        sug     = [classes[i] for i in idx]
        conf    = [float(proba[r, i]) for r, i in enumerate(idx)]
    else:
        sug  = list(pipe.predict(X))
        conf = [1.0] * len(sug)
    return sug, conf


# ── clasificación principal ───────────────────────────────────────────────────

def classify(excel_path=EXCEL_PATH, sheet_name=SHEET_CA, write_back=True, verbose=True):
    if not HAS_SKLEARN:
        print("⚠ scikit-learn no instalado. Instalá con: pip install scikit-learn openpyxl")
        return None

    df = load_sheet(excel_path, sheet_name)

    excluir = EXCLUIR_DESC | {"Traspaso", "Saldo Inicial", "Saldo Final", "No va"}
    train_df = df[~df["Descripción"].isin(excluir) & df["Descripción"].ne("")]

    if train_df["Descripción"].nunique() < 2:
        print("No hay suficientes categorías etiquetadas (mínimo 2).")
        return df

    if verbose:
        print(f"Entrenando con {len(train_df)} filas y {train_df['Descripción'].nunique()} categorías...")

    pipe = build_pipeline(train_df)

    # filas sin clasificar
    desc = df["Descripción"]
    pred_mask = desc.eq("") | desc.str.upper().isin(["NAN", "NONE", "SIN CLASIFICAR", "PENDIENTE"])
    n_pred = pred_mask.sum()

    if verbose:
        print(f"Filas a clasificar: {n_pred}")

    if n_pred > 0:
        X = df.loc[pred_mask, ["TextoML", "Monto"]].copy()
        sug, conf = predict_proba_safe(pipe, X)

        df.loc[pred_mask, "Descripción"]  = sug
        df.loc[pred_mask, "ML_Confianza"] = conf
        df.loc[pred_mask, "ML_Revisar"]   = [c < CONFIANZA_MINIMA for c in conf]
        n_revisar = sum(c < CONFIANZA_MINIMA for c in conf)
        if verbose:
            print(f"  ✓ clasificadas automáticamente: {n_pred - n_revisar}")
            print(f"  ⚠ para revisar (confianza baja): {n_revisar}")

    if write_back and n_pred > 0:
        _write_back(df, excel_path, sheet_name, verbose)

    return df


def _write_back(df, excel_path, sheet_name, verbose=True):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    desc_col  = headers.get("Descripción")
    desc2_col = headers.get("Descripción 2")

    # Agregar columnas ML si no existen
    next_col = ws.max_column + 1
    if "ML_Revisar" not in headers:
        headers["ML_Revisar"] = next_col
        ws.cell(1, next_col).value = "ML_Revisar"
        next_col += 1
    if "ML_Confianza" not in headers:
        headers["ML_Confianza"] = next_col
        ws.cell(1, next_col).value = "ML_Confianza"

    revisar_col  = headers["ML_Revisar"]
    confianza_col = headers["ML_Confianza"]

    yellow_fill = PatternFill("solid", fgColor="FFF59D")
    red_font    = Font(color="C62828")

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        if desc_col:
            ws.cell(row_idx, desc_col).value = row["Descripción"]
        if desc2_col:
            ws.cell(row_idx, desc2_col).value = row["Descripción 2"]

        revisar = bool(row.get("ML_Revisar", False))
        conf_val = row.get("ML_Confianza")

        ws.cell(row_idx, revisar_col).value = revisar
        ws.cell(row_idx, confianza_col).value = (
            None if pd.isna(conf_val) else round(float(conf_val), 4)
        )
        if revisar and desc_col:
            ws.cell(row_idx, desc_col).fill = yellow_fill
            ws.cell(row_idx, desc_col).font = red_font

    wb.save(excel_path)
    if verbose:
        print(f"✓ Excel actualizado: {excel_path}")


if __name__ == "__main__":
    classify()
