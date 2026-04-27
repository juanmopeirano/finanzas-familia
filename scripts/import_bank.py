"""
import_bank.py — agrega movimientos nuevos del banco al Excel maestro.

Uso:
  python import_bank.py ruta/al/archivo.csv
  python import_bank.py ruta/al/archivo.xlsx
  python import_bank.py ruta/al/archivo.xlsx --sheet "Nombre hoja"

Itaú exporta estados de cuenta en Excel o CSV.
El script detecta las columnas automáticamente y agrega solo las filas nuevas
(deduplica por Fecha + Concepto + Débito + Crédito).
"""

import sys
import re
import hashlib
import argparse
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
from config import EXCEL_PATH, SHEET_CA
from reglas import clasificar, cargar_reglas

# Categorías de ingreso "indudables" (sueldo, etc).
# Crédito de otra categoría (REDIVA, reintegros, etc.) NO es ingreso → es gasto neteado.
INGRESO_CATS = {"Sueldo JM", "Sueldo Pili", "Otros ingresos", "Ahorros"}
SISTEMA_CATS = {"No va", "Traspaso"}


# ── normalización de nombres de columna ──────────────────────────────────────

_COL_ALIASES = {
    "fecha":      ["fecha", "date", "fec"],
    "concepto":   ["concepto", "descripcion", "descripción", "detalle", "glosa", "movimiento"],
    "debito":     ["debito", "débito", "cargo", "egreso", "monto débito", "monto debito"],
    "credito":    ["credito", "crédito", "abono", "ingreso", "monto crédito", "monto credito"],
    "saldo":      ["saldo", "balance", "saldo final"],
    "referencia": ["referencia", "ref", "num. operacion", "nro operacion", "operacion"],
}


def _normalize_col(name):
    """Normaliza nombre de columna para comparar."""
    s = str(name).lower().strip()
    s = s.replace("á","a").replace("é","e").replace("í","i").replace("ó","o").replace("ú","u")
    return re.sub(r"[^a-z0-9]", "", s)


def _match_columns(raw_cols):
    """Devuelve dict {campo_estandar: nombre_real} para las columnas encontradas."""
    norm_map = {_normalize_col(c): c for c in raw_cols}
    result = {}
    for field, aliases in _COL_ALIASES.items():
        for alias in aliases:
            key = _normalize_col(alias)
            if key in norm_map:
                result[field] = norm_map[key]
                break
    return result


# ── lectura del archivo bancario ──────────────────────────────────────────────

def read_bank_file(path, sheet=None):
    path = Path(path)
    ext  = path.suffix.lower()

    # intentar leer saltando filas vacías al principio
    if ext in (".xlsx", ".xls", ".xlsm"):
        raw = pd.read_excel(path, sheet_name=(sheet or 0), header=None, dtype=str)
    else:
        # CSV: probar distintos separadores
        for sep in [",", ";", "\t", "|"]:
            try:
                raw = pd.read_csv(path, sep=sep, header=None, dtype=str, encoding="utf-8")
                if raw.shape[1] > 2:
                    break
            except Exception:
                continue
        else:
            raw = pd.read_csv(path, header=None, dtype=str, encoding="latin-1")

    # encontrar la fila de encabezados (la primera que tenga "fecha" o "concepto")
    header_row = 0
    for i, row in raw.iterrows():
        vals = " ".join(str(v).lower() for v in row.dropna())
        if any(k in vals for k in ["fecha", "concepto", "debito", "débito", "credito", "crédito"]):
            header_row = i
            break

    raw.columns = raw.iloc[header_row]
    df = raw.iloc[header_row + 1:].reset_index(drop=True)
    df = df.dropna(how="all")

    col_map = _match_columns(df.columns.tolist())
    if "fecha" not in col_map or ("debito" not in col_map and "credito" not in col_map):
        raise ValueError(
            f"No se encontraron columnas de fecha/débito/crédito en el archivo.\n"
            f"Columnas detectadas: {list(df.columns)}"
        )

    # renombrar al esquema estándar
    rename = {v: k for k, v in col_map.items()}
    df = df.rename(columns=rename)

    # parsear tipos
    df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce", dayfirst=True)
    for col in ["debito", "credito", "saldo"]:
        if col in df.columns:
            df[col] = (
                df[col].astype(str)
                .str.replace(r"[^\d,\.]", "", regex=True)
                .str.replace(",", ".")
                .replace("", "0")
                .astype(float)
                .fillna(0.0)
            )
        else:
            df[col] = 0.0

    df = df.dropna(subset=["fecha"])

    # Filtrar filas tipo "SALDO ANTERIOR" / "SALDO FINAL" / "TOTAL" / etc.
    if "concepto" in df.columns:
        excluir = ["SALDO ANTERIOR", "SALDO FINAL", "SALDO INICIAL", "TOTAL"]
        df = df[~df["concepto"].astype(str).str.upper().str.strip().isin(excluir)]

    return df.reset_index(drop=True)


# ── deduplicación ─────────────────────────────────────────────────────────────

def _row_key(fecha, concepto, debito, credito):
    """Clave única para deduplicar movimientos. Trata NaN y 0 como equivalentes."""
    def num(x):
        try:
            v = float(x)
            return 0.0 if pd.isna(v) else round(v, 2)
        except Exception:
            return 0.0
    f = str(fecha)[:10]  # YYYY-MM-DD
    c = str(concepto).strip().upper() if concepto is not None else ""
    raw = f"{f}|{c}|{num(debito)}|{num(credito)}"
    return hashlib.md5(raw.encode()).hexdigest()[:12]


def build_existing_keys(ws_df):
    keys = set()
    for _, r in ws_df.iterrows():
        k = _row_key(
            r.get("Fecha", ""),
            r.get("Concepto", ""),
            r.get("Débito", 0),
            r.get("Crédito", 0),
        )
        keys.add(k)
    return keys


# ── escritura al Excel maestro ────────────────────────────────────────────────

def _is_color(cell, hex6):
    """Devuelve True si la celda tiene fondo del color indicado."""
    try:
        if cell.fill is None or cell.fill.fgColor is None: return False
        v = cell.fill.fgColor.value
        if v is None: return False
        return str(v).upper().endswith(hex6.upper())
    except Exception:
        return False


def append_to_master(new_rows: pd.DataFrame, excel_path=EXCEL_PATH, sheet_name=SHEET_CA, origen="Caja de Ahorro"):
    """
    Agrega filas nuevas al Excel maestro, auto-clasificando con reglas.
    Pinta las filas nuevas en celeste; las nuevas-de-runs-anteriores se despintan.
    Categoría/Subcategoría quedan amarillas si hay 'Sin clasificar'.
    Devuelve (added, sin_clasificar).
    """
    from openpyxl.styles import PatternFill
    cargar_reglas()

    LIGHT_BLUE = "DCEBFF"   # filas nuevas de este import
    YELLOW     = "FEF3C7"   # categoría sin clasificar

    wb   = load_workbook(excel_path)
    ws   = wb[sheet_name]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(headers) if h}

    # 1) Limpiar el celeste de imports previos (mantiene amarillo)
    blue_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    no_fill   = PatternFill(fill_type=None)
    for row_num in range(2, ws.max_row + 1):
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row_num, col_num)
            if _is_color(cell, LIGHT_BLUE):
                cell.fill = no_fill

    next_row = ws.max_row + 1
    yellow_fill = PatternFill("solid", fgColor=YELLOW)
    added    = 0
    sin_clas = 0

    for _, r in new_rows.iterrows():
        concepto = str(r.get("concepto", "") or "").strip()
        cat, sub = clasificar(concepto)
        es_sin = (cat == "Sin clasificar")
        sin_clas += int(es_sin)

        # Débito/Crédito limpios (None si vacíos)
        debito  = r["debito"]  if r["debito"]  > 0 else None
        credito = r["credito"] if r["credito"] > 0 else None

        # Tipo automático según CATEGORÍA (no según columna del banco)
        # Un crédito puede ser un reintegro/REDIVA que netea un gasto, NO un ingreso.
        if cat == "Sin clasificar":
            tipo = ""                # vos decidís cuando lo revises
        elif cat in INGRESO_CATS:
            tipo = "Ingresos"
        elif cat in SISTEMA_CATS:
            tipo = ""                # No va / Traspaso
        else:
            tipo = "Gastos"          # default: incluye créditos que netean gastos

        row_data = {
            "Tipo":          tipo,
            "Origen":        origen,
            "Categoría":     cat,
            "Subcategoría":  sub,
            "Fecha":         r["fecha"].date() if pd.notna(r["fecha"]) else "",
            "Concepto":      concepto,
            "Débito":        debito,
            "Crédito":       credito,
            "Saldo":         r.get("saldo") or None,
            "Referencia":    r.get("referencia", "") or "",
            "Destino":       "",
            "ML_Revisar":    es_sin,
        }
        for field, col_num in col_idx.items():
            if field in row_data:
                cell = ws.cell(next_row, col_num)
                cell.value = row_data[field]
                # Por defecto: celeste (fila nueva en este import)
                cell.fill = blue_fill
                # Override: amarillo en Categoría/Subcategoría si requiere revisión
                if es_sin and field in ("Categoría", "Subcategoría"):
                    cell.fill = yellow_fill
        next_row += 1
        added += 1

    wb.save(excel_path)
    return added, sin_clas


# ── función principal ─────────────────────────────────────────────────────────

def import_bank(bank_file, sheet=None, excel_path=EXCEL_PATH, master_sheet=SHEET_CA, verbose=True):
    bank_file = Path(bank_file)
    if not bank_file.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {bank_file}")

    if verbose:
        print(f"Leyendo: {bank_file.name}")

    new_df = read_bank_file(bank_file, sheet)
    if verbose:
        print(f"  {len(new_df)} movimientos en el archivo del banco")

    # cargar existentes para deduplicar
    existing = pd.read_excel(excel_path, sheet_name=master_sheet, engine="openpyxl")
    existing_keys = build_existing_keys(existing)

    new_df["_key"] = new_df.apply(
        lambda r: _row_key(r["fecha"], r.get("concepto", ""), r["debito"], r["credito"]),
        axis=1
    )
    to_add = new_df[~new_df["_key"].isin(existing_keys)].drop(columns=["_key"])

    if len(to_add) == 0:
        if verbose:
            print("  ya estaban todos importados (0 nuevos)")
        return 0, 0

    added, sin_clas = append_to_master(to_add, excel_path, master_sheet)
    if verbose:
        print(f"  + {added} movimientos nuevos agregados")
        print(f"  - {added - sin_clas} clasificados automaticamente")
        if sin_clas:
            print(f"  ! {sin_clas} requieren revision manual (amarillo en Excel)")
    return added, sin_clas


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Importar estado de cuenta bancario")
    parser.add_argument("archivo", help="Ruta al CSV o Excel del banco")
    parser.add_argument("--sheet", default=None, help="Nombre de hoja (solo para Excel)")
    args = parser.parse_args()
    import_bank(args.archivo, sheet=args.sheet)
