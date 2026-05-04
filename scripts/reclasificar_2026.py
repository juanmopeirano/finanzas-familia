"""
reclasificar_2026.py — script ONE-SHOT.

Aplica las reglas actuales de mapping_propuesto.xlsx a TODOS los movimientos
del año 2026 en adelante en Planilla Familia.xlsx (hoja CA).

Reglas:
  - Si una regla matchea: actualiza Categoría / Subcategoría / Origen / Tipo
    (Origen y Tipo solo si la regla los tiene definidos)
  - Si NO matchea (Sin clasificar): deja la fila como está, no toca nada
  - Filas anteriores a 2026: NO se tocan (histórico viejo intacto)

Uso:
  python reclasificar_2026.py            # DRY-RUN: muestra qué cambiaría sin guardar
  python reclasificar_2026.py --apply    # aplica los cambios y guarda

Recomiendo: corré el dry-run primero, mirá los stats, y después --apply.
"""

import sys
import os
import shutil
import argparse
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
from config import EXCEL_PATH, SHEET_CA
from reglas import clasificar, cargar_reglas


ANO_DESDE = 2026


def _to_float_safe(v):
    if v is None: return 0.0
    try: return float(v)
    except Exception: return 0.0


def _to_str_safe(v):
    if v is None: return ""
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none") else s


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true",
                        help="Aplica los cambios. Sin esta flag corre en dry-run.")
    args = parser.parse_args()
    dry_run = not args.apply

    print("=" * 60)
    print(f"RE-CLASIFICAR MOVS {ANO_DESDE}+")
    if dry_run:
        print(f"MODO: dry-run (no se guarda nada)")
    else:
        print(f"MODO: APPLY (se va a guardar)")
    print("=" * 60)

    print(f"\nCargando reglas...")
    reglas = cargar_reglas(force=True)
    print(f"  {len(reglas)} reglas activas")

    print(f"\nAbriendo {EXCEL_PATH.name}...")
    try:
        wb = load_workbook(EXCEL_PATH)
    except PermissionError:
        print("ERROR: el archivo está abierto en Excel. Cerralo y volvé a correr.")
        sys.exit(1)

    if SHEET_CA not in wb.sheetnames:
        print(f"ERROR: no se encuentra la hoja '{SHEET_CA}'")
        sys.exit(1)

    ws = wb[SHEET_CA]

    # Mapear columnas por nombre
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    cols_required = ["Fecha", "Concepto", "Categoría"]
    for cn in cols_required:
        if cn not in headers:
            print(f"ERROR: falta la columna '{cn}'")
            sys.exit(1)

    col_fecha    = headers["Fecha"]
    col_concepto = headers["Concepto"]
    col_debito   = headers.get("Débito")
    col_credito  = headers.get("Crédito")
    col_cat      = headers["Categoría"]
    col_sub      = headers.get("Subcategoría")
    col_origen   = headers.get("Origen")
    col_tipo     = headers.get("Tipo")

    print(f"\nProcesando filas...")

    total_2026     = 0
    matched        = 0
    unchanged      = 0
    cat_changed    = 0
    sub_changed    = 0
    origen_changed = 0
    tipo_changed   = 0

    # Ejemplos para mostrar
    ejemplos_cambio = []

    for row_idx in range(2, ws.max_row + 1):
        fecha_val = ws.cell(row_idx, col_fecha).value
        if fecha_val is None: continue

        # Parsear fecha (puede ser datetime/date/string)
        try:
            year = fecha_val.year if hasattr(fecha_val, "year") \
                   else pd.to_datetime(fecha_val, dayfirst=True, errors="coerce").year
        except Exception:
            continue
        if not year or year < ANO_DESDE:
            continue

        total_2026 += 1

        concepto = _to_str_safe(ws.cell(row_idx, col_concepto).value)
        debito   = _to_float_safe(ws.cell(row_idx, col_debito).value)   if col_debito  else 0
        credito  = _to_float_safe(ws.cell(row_idx, col_credito).value)  if col_credito else 0

        new_cat, new_sub, new_origen, new_tipo = clasificar(concepto, debito, credito)

        if new_cat == "Sin clasificar":
            unchanged += 1
            continue

        matched += 1

        # Estados actuales
        old_cat    = _to_str_safe(ws.cell(row_idx, col_cat).value)
        old_sub    = _to_str_safe(ws.cell(row_idx, col_sub).value)    if col_sub    else ""
        old_origen = _to_str_safe(ws.cell(row_idx, col_origen).value) if col_origen else ""
        old_tipo   = _to_str_safe(ws.cell(row_idx, col_tipo).value)   if col_tipo   else ""

        cambios_fila = []

        # Categoría: siempre se actualiza si difiere
        if old_cat != new_cat:
            if not dry_run: ws.cell(row_idx, col_cat).value = new_cat
            cat_changed += 1
            cambios_fila.append(f"cat: '{old_cat}' -> '{new_cat}'")

        # Subcategoría: solo si la regla la tiene Y difiere
        if new_sub and col_sub and old_sub != new_sub:
            if not dry_run: ws.cell(row_idx, col_sub).value = new_sub
            sub_changed += 1
            cambios_fila.append(f"sub: '{old_sub}' -> '{new_sub}'")

        # Origen: solo si la regla lo tiene Y difiere
        if new_origen and col_origen and old_origen != new_origen:
            if not dry_run: ws.cell(row_idx, col_origen).value = new_origen
            origen_changed += 1
            cambios_fila.append(f"origen: '{old_origen}' -> '{new_origen}'")

        # Tipo: solo si la regla lo tiene Y difiere
        if new_tipo and col_tipo and old_tipo != new_tipo:
            if not dry_run: ws.cell(row_idx, col_tipo).value = new_tipo
            tipo_changed += 1
            cambios_fila.append(f"tipo: '{old_tipo}' -> '{new_tipo}'")

        if cambios_fila and len(ejemplos_cambio) < 15:
            fecha_str = fecha_val.strftime("%d/%m/%Y") if hasattr(fecha_val, "strftime") else str(fecha_val)
            ejemplos_cambio.append((fecha_str, concepto[:35], cambios_fila))

    print(f"\n=== Resumen ===")
    print(f"Total filas {ANO_DESDE}+:                {total_2026}")
    print(f"  Matchearon alguna regla:           {matched}")
    print(f"  Sin clasificar (intactas):         {unchanged}")
    print(f"\nCambios {'(simulados)' if dry_run else 'APLICADOS'}:")
    print(f"  Categoría modificada:              {cat_changed}")
    print(f"  Subcategoría modificada:           {sub_changed}")
    print(f"  Origen modificado:                 {origen_changed}")
    print(f"  Tipo modificado:                   {tipo_changed}")

    if ejemplos_cambio:
        print(f"\nEjemplos de cambios (primeros {len(ejemplos_cambio)}):")
        print("-" * 100)
        for fecha_s, conc, cambios in ejemplos_cambio:
            print(f"  {fecha_s}  {conc:35}")
            for c in cambios:
                print(f"               | {c}")

    if dry_run:
        print(f"\n[DRY-RUN] Nada se guardó. Para aplicar:")
        print(f"  python reclasificar_2026.py --apply")
    else:
        # Backup antes de guardar
        backup = EXCEL_PATH.parent / f"{EXCEL_PATH.stem}_BACKUP_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        try:
            shutil.copy2(EXCEL_PATH, backup)
            print(f"\nBackup creado: {backup.name}")
        except Exception as e:
            print(f"AVISO: no pude crear backup: {e}")

        print(f"Guardando cambios...")
        wb.save(EXCEL_PATH)
        print(f"OK - Guardado: {EXCEL_PATH.name}")


if __name__ == "__main__":
    main()
