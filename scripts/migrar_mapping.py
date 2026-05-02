"""
migrar_mapping.py — script ONE-SHOT para migrar el formato viejo de
mapping_propuesto.xlsx (hojas 'Patrones' + 'Sin patron') al nuevo formato
unificado en una sola hoja 'Reglas'.

Nuevo formato:
  - Patrón texto · Monto desde · Monto hasta · Categoría · Subcategoría

Solo se ejecuta una vez. Después se edita la hoja 'Reglas' a mano en Excel.
"""

import os
import sys
import shutil
import tempfile
import warnings
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).parent))
from config import PROJECT_DIR
from generar_mapping import CATEGORIAS

warnings.filterwarnings("ignore", message="Data Validation extension is not supported")

MAPPING_FILE = PROJECT_DIR / "mapping_propuesto.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="064E3B")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN        = Side(border_style="thin", color="E5E7EB")
BOX         = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)


def _read_safe(path, sheet):
    """Lee la hoja salvando bloqueo de OneDrive si pasa."""
    try:
        return pd.read_excel(path, sheet_name=sheet, engine="openpyxl")
    except PermissionError:
        tmp = os.path.join(tempfile.gettempdir(), f"_mig_{Path(path).name}")
        shutil.copy2(path, tmp)
        return pd.read_excel(tmp, sheet_name=sheet, engine="openpyxl")


# ── Migración ────────────────────────────────────────────────────────────────

def construir_reglas(wb):
    """Lee Patrones + Sin patrón (formato viejo) o Reglas (formato nuevo)
    y devuelve lista unificada de reglas."""
    reglas = []
    sheets = wb.sheetnames

    def _safe_str(v):
        if v is None or pd.isna(v): return ""
        s = str(v).strip()
        return "" if s.lower() in ("nan", "none") else s

    # Si ya está en formato nuevo, leer la hoja Reglas existente
    if "Reglas" in sheets and "Patrones" not in sheets and "Sin patrón" not in sheets:
        print("  Detectado formato NUEVO: leyendo hoja 'Reglas'...")
        df_r = _read_safe(MAPPING_FILE, "Reglas")
        df_r = df_r[df_r["Categoría"].notna()].copy()
        for _, r in df_r.iterrows():
            reglas.append({
                "patron": _safe_str(r.get("Patrón texto")),
                "desde": float(r["Monto desde"]) if pd.notna(r.get("Monto desde")) else None,
                "hasta": float(r["Monto hasta"]) if pd.notna(r.get("Monto hasta")) else None,
                "cat":   _safe_str(r.get("Categoría")),
                "sub":   _safe_str(r.get("Subcategoría")),
                "origen": _safe_str(r.get("Origen")),
                "tipo":   _safe_str(r.get("Tipo")),
            })
        print(f"  + {len(reglas)} reglas leídas")
        return reglas

    # Formato viejo: combinar Patrones + Sin patrón
    print("  Detectado formato VIEJO: combinando 'Patrones' + 'Sin patrón'...")
    if "Patrones" in sheets:
        df_p = _read_safe(MAPPING_FILE, "Patrones")
        df_p = df_p[df_p["Categoría"].notna()].copy()
        for _, r in df_p.iterrows():
            reglas.append({
                "patron": _safe_str(r.get("Patrón (token banco)")),
                "desde": None,
                "hasta": None,
                "cat":   _safe_str(r.get("Categoría")),
                "sub":   _safe_str(r.get("Subcategoría")),
                "origen": "",
                "tipo":   "",
            })
        print(f"  + {len(df_p)} reglas desde 'Patrones'")

    if "Sin patrón" in sheets:
        df_s = _read_safe(MAPPING_FILE, "Sin patrón")
        df_s = df_s[df_s["Categoría a asignar"].notna()].copy()
        for _, r in df_s.iterrows():
            reglas.append({
                "patron": _safe_str(r.get("Concepto normalizado")),
                "desde": None,
                "hasta": None,
                "cat":   _safe_str(r.get("Categoría a asignar")),
                "sub":   _safe_str(r.get("Subcategoría")),
                "origen": "",
                "tipo":   "",
            })
        print(f"  + {len(df_s)} reglas desde 'Sin patrón'")

    # Ordenar para presentación humana
    reglas.sort(key=lambda r: (r["cat"], -len(r["patron"]), r["patron"]))
    return reglas


def asegurar_columna_subcategorias(wb):
    """Pobla la columna D de la hoja Categorías con todas las subcategorías
    (necesario para el dropdown de Subcategoría en la hoja Reglas)."""
    ws = wb["Categorías"]
    # Lista plana de todas las subcategorías
    subs_flat = sorted(set(s for subs in CATEGORIAS.values() for s in subs))

    # Header en D1 si no existe
    if not ws.cell(1, 4).value:
        cell = ws.cell(1, 4, "Subcategorías")
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    # Limpiar columna D existente (de fila 2 hacia abajo)
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 4).value = None

    # Escribir lista
    for i, s in enumerate(subs_flat, 2):
        ws.cell(i, 4, s)

    ws.column_dimensions["D"].width = 24
    return len(subs_flat)


def contar_categorias(wb):
    """Cuenta cuántas categorías (col A) hay en la hoja Categorías."""
    ws = wb["Categorías"]
    n = 0
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value:
            n += 1
    return n


def crear_hoja_reglas(wb, reglas, n_cats, n_subs):
    """Crea (o reemplaza) la hoja 'Reglas' en el workbook."""
    if "Reglas" in wb.sheetnames:
        del wb["Reglas"]

    ws = wb.create_sheet("Reglas", 1)  # segunda posición (después de Instrucciones)

    headers = ["Patrón texto", "Monto desde", "Monto hasta",
               "Categoría", "Subcategoría", "Origen", "Tipo"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Data
    for i, r in enumerate(reglas, 2):
        ws.cell(i, 1, r["patron"]).alignment = Alignment(vertical="center")
        if r["desde"] is not None:
            ws.cell(i, 2, r["desde"])
        if r["hasta"] is not None:
            ws.cell(i, 3, r["hasta"])
        ws.cell(i, 4, r["cat"]).font = Font(bold=True)
        ws.cell(i, 5, r["sub"])
        if r.get("origen"): ws.cell(i, 6, r["origen"])
        if r.get("tipo"):   ws.cell(i, 7, r["tipo"])
        for c in range(1, 8):
            ws.cell(i, c).border = BOX

    # Anchos
    widths = [40, 14, 14, 24, 22, 18, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Formato numérico para columnas de monto
    for i in range(2, len(reglas) + 2):
        ws.cell(i, 2).number_format = "#,##0.00"
        ws.cell(i, 3).number_format = "#,##0.00"

    # Freeze + autofilter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(reglas) + 1}"

    # Data validation: dropdowns
    range_max = max(len(reglas) + 200, 1000)  # margen para futuras

    dv_cat = DataValidation(type="list",
                             formula1=f"=Categorías!$A$2:$A${n_cats + 1}",
                             allow_blank=True)
    dv_cat.error = "Elegí una categoría de la lista de la hoja Categorías"
    dv_cat.errorTitle = "Categoría inválida"
    ws.add_data_validation(dv_cat)
    dv_cat.add(f"D2:D{range_max}")

    dv_sub = DataValidation(type="list",
                             formula1=f"=Categorías!$D$2:$D${n_subs + 1}",
                             allow_blank=True)
    ws.add_data_validation(dv_sub)
    dv_sub.add(f"E2:E{range_max}")

    dv_origen = DataValidation(type="list",
                               formula1='"Caja de Ahorro,Tarjeta de crédito"',
                               allow_blank=True)
    ws.add_data_validation(dv_origen)
    dv_origen.add(f"F2:F{range_max}")

    dv_tipo = DataValidation(type="list",
                             formula1='"Ingresos,Gastos"',
                             allow_blank=True)
    ws.add_data_validation(dv_tipo)
    dv_tipo.add(f"G2:G{range_max}")

    return ws


def actualizar_instrucciones(wb):
    """Reescribe la hoja 'Instrucciones' con el nuevo formato."""
    if "Instrucciones" in wb.sheetnames:
        del wb["Instrucciones"]
    ws = wb.create_sheet("Instrucciones", 0)  # primera posición

    instr = [
        ("FINANZAS — MAPPING DE REGLAS", True),
        ("", False),
        ("Esta planilla define las reglas de auto-clasificación de movimientos.", False),
        ("", False),
        ("HOJAS:", True),
        ("  • Reglas — todas las reglas en una sola tabla", False),
        ("  • Categorías — referencia de categorías y subcategorías válidas", False),
        ("", False),
        ("HOJA 'Reglas' — cómo agregar/editar:", True),
        ("", False),
        ("  • Patrón texto: parte de la descripción del banco a buscar", False),
        ("    (ej 'UTE', 'TIENDA INGLE'). Case-insensitive (mayús/minús da igual).", False),
        ("    Vacío = no se chequea texto.", False),
        ("", False),
        ("  • Monto desde / Monto hasta: rango opcional. Se compara contra", False),
        ("    Débito o Crédito (el que tenga valor). Vacíos = no se chequea monto.", False),
        ("", False),
        ("  • Una regla matchea si TODAS sus condiciones (las que están completas)", False),
        ("    se cumplen.", False),
        ("", False),
        ("  • Si varias reglas matchean: gana la que tiene MÁS condiciones,", False),
        ("    después la de patrón más largo, después la primera de la hoja.", False),
        ("", False),
        ("EJEMPLOS:", True),
        ("", False),
        ("  • Patrón='UTE', Categoría='Servicios', Subcategoría='UTE'", False),
        ("    → cualquier mov con 'UTE' en el concepto va a Servicios/UTE", False),
        ("", False),
        ("  • Patrón='TRASPASO DE 7765226', Monto desde=5000, Monto hasta=200000,", False),
        ("    Categoría='Sueldo Pili'", False),
        ("    → solo los traspasos de esa cuenta entre 5k y 200k son Sueldo Pili", False),
        ("    (los más chicos del mismo concepto NO se clasifican como sueldo)", False),
        ("", False),
        ("  • Patrón vacío, Monto desde=0, Monto hasta=10, Categoría='Varios'", False),
        ("    → cualquier mov pequeño entre 0 y 10 va a Varios", False),
        ("", False),
        ("DESPUÉS DE EDITAR:", True),
        ("  Guardá el Excel y volvé a correr 'Actualizar Datos.bat' o 'Publicar Todo.bat'", False),
        ("  para que las reglas se apliquen a los movimientos.", False),
    ]
    for r, (txt, is_header) in enumerate(instr, 1):
        cell = ws.cell(r, 1, txt)
        if is_header:
            cell.font = Font(bold=True, size=13, color="064E3B")
        else:
            cell.font = Font(size=11, color="1F2937")
    ws.column_dimensions["A"].width = 90


def main():
    if not MAPPING_FILE.exists():
        print(f"ERROR: no se encuentra {MAPPING_FILE}")
        sys.exit(1)

    print(f"Leyendo {MAPPING_FILE.name}...")
    try:
        wb = load_workbook(MAPPING_FILE)
    except PermissionError:
        print("ERROR: el archivo está abierto en Excel. Cerralo y volvé a correr.")
        sys.exit(1)

    reglas = construir_reglas(wb)
    print(f"  Total: {len(reglas)} reglas")

    print("\nReconstruyendo hojas...")
    n_subs = asegurar_columna_subcategorias(wb)
    n_cats = contar_categorias(wb)
    print(f"  Categorías: {n_cats}, subcategorías para dropdown: {n_subs}")

    crear_hoja_reglas(wb, reglas, n_cats, n_subs)
    actualizar_instrucciones(wb)

    # Eliminar hojas viejas (si quedaran)
    for old_sheet in ("Patrones", "Sin patrón"):
        if old_sheet in wb.sheetnames:
            del wb[old_sheet]
            print(f"  - Hoja '{old_sheet}' eliminada")

    wb.save(MAPPING_FILE)
    try:
        print(f"\nOK - {len(reglas)} reglas migradas. Hojas viejas eliminadas.")
        print(f"Archivo: {MAPPING_FILE}")
    except UnicodeEncodeError:
        print(f"\nOK ({len(reglas)} reglas migradas)")


if __name__ == "__main__":
    main()
